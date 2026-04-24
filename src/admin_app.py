from __future__ import annotations

import hashlib
import io
import json
import mimetypes
import os
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Annotated, Any, Optional

import jwt
import pymysql
from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from .db import db, normalize_target_asin, parse_task_params
from .totp_qr import totp_current_code
from .task_report_parse import parse_task_report_footer
from .paths import STATIC_WEB_DIR, TASK_IMAGE_DIR, ensure_data_dir, safe_task_image_path

STATIC = STATIC_WEB_DIR

JWT_ALG = "HS256"
JWT_EXPIRE_DAYS = 7


def _jwt_signing_key() -> bytes:
    raw = os.getenv("JWT_SECRET", "dev-change-JWT_SECRET-in-production").encode("utf-8")
    if len(raw) < 32:
        return hashlib.sha256(raw).digest()
    return raw


oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/v1/auth/token")


def _create_token(user_id: int, username: str, is_admin: bool) -> str:
    now = datetime.now(timezone.utc)
    payload = {
        "sub": str(user_id),
        "username": username,
        "is_admin": is_admin,
        "iat": now,
        "exp": now + timedelta(days=JWT_EXPIRE_DAYS),
    }
    return jwt.encode(payload, _jwt_signing_key(), algorithm=JWT_ALG)


def _decode_token(token: str) -> dict[str, Any]:
    try:
        return jwt.decode(token, _jwt_signing_key(), algorithms=[JWT_ALG])
    except jwt.InvalidTokenError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="鏃犳晥鎴栬繃鏈熺殑浠ょ墝",
            headers={"WWW-Authenticate": "Bearer"},
        )


async def get_current_user(token: Annotated[str, Depends(oauth2_scheme)]) -> dict[str, Any]:
    data = _decode_token(token)
    uid = data.get("sub")
    if not uid:
        raise HTTPException(status_code=401, detail="鏃犳晥浠ょ墝")
    return {
        "id": int(uid),
        "username": data.get("username", ""),
        "is_admin": bool(data.get("is_admin", False)),
    }


CurrentUser = Annotated[dict[str, Any], Depends(get_current_user)]


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"


class MeResponse(BaseModel):
    id: int
    username: str
    is_admin: bool


class PaginatedTasks(BaseModel):
    items: list[dict[str, Any]]
    page: int
    per_page: int
    total: int
    total_pages: int


class PaginatedRows(BaseModel):
    items: list[dict[str, Any]]
    page: int
    per_page: int
    total: int
    total_pages: int


class DeviceAliasBody(BaseModel):
    alias: str | None = None


class DeviceScreenshotPolicyBody(BaseModel):
    screenshot_upload_policy: str = Field(..., pattern="^(all|failed_only|none)$")


class BatchClickTasksBody(BaseModel):
    task_type: str = Field(..., min_length=1)
    keyword: str = Field(..., min_length=1)
    res_folder_name: str = Field(
        "",
        min_length=0,
        description="涓庡叧閿瘝 1:1锛涘鎴风 res 涓嬭祫婧愮洰褰曞悕",
    )
    product_title: str = Field(
        "",
        min_length=0,
        description="deprecated, for backward compatibility only",
    )
    product_titles: list[str] = Field(
        default_factory=list,
        description="deprecated, for backward compatibility only",
    )
    mode: str = Field(..., pattern="^(manual|smart)$")
    device_ids: list[str] = Field(default_factory=list)
    per_device_counts: dict[str, int] = Field(default_factory=dict)
    total_count: int = Field(0, ge=0, le=100000)
    save_data_record: bool = Field(
        False, description="鍕鹃€夊悗浠诲姟鎴愬姛缁撴鏃跺啓鍏ュ綊妗ｏ紙浠呮垚鍔燂紝鍚弬鏁颁笌鐘舵€侊級"
    )


class BatchRegisterBody(BaseModel):
    mode: str = Field(..., pattern="^(manual|smart)$")
    device_ids: list[str] = Field(default_factory=list)
    per_device_counts: dict[str, int] = Field(default_factory=dict)
    total_count: int = Field(0, ge=0, le=100000, description="鏅鸿兘鍒嗛厤鏃讹細瑕佸垱寤虹殑娉ㄥ唽浠诲姟鎬绘暟")
    bind_email: bool = Field(
        False, description="涓烘瘡涓换鍔′粠閭鎺ョ爜搴撶粦瀹氫竴鏉★紙涓庢墜鏈轰竴涓€瀵瑰簲锛屽悇搴撳潎浠呰兘鐢ㄤ竴娆★級"
    )
    save_data_record: bool = Field(
        True,
        description="default true: save creation snapshot and merge final result on task completion",
    )


class AdminSettingsResponse(BaseModel):
    task_retention_days: int


class AdminSettingsPatchBody(BaseModel):
    task_retention_days: int = Field(..., ge=1, le=3650)


class UsAddressBody(BaseModel):
    recipient_name: str | None = None
    state: str | None = None
    city: str | None = None
    address_line1: str | None = None
    address_line2: str | None = None
    zip_code: str | None = None
    phone: str | None = None
    full_line: str | None = None


class TargetAsinCreateBody(BaseModel):
    asin: str = Field(..., min_length=1, max_length=32)
    note: str = Field("", max_length=512)


class TargetAsinPatchBody(BaseModel):
    asin: str | None = Field(None, min_length=1, max_length=32)
    note: str | None = Field(None, max_length=512)


class HeartbeatBody(BaseModel):
    device_id: str = Field(..., min_length=1, max_length=128)


class ClientAsinClickBody(BaseModel):
    device_id: str = Field(..., min_length=1, max_length=128)
    asin: str = Field(..., min_length=1, max_length=32)
    keyword: str = Field(..., min_length=1, max_length=512)


class ClientAmazonBootstrapBody(BaseModel):
    device_id: str = Field(..., min_length=1, max_length=128)
    task_id: int = Field(..., ge=1)


class ClientAmazonAccountStageBody(BaseModel):
    phone: str = Field(..., min_length=1, max_length=64)


class TaskReportParsePreviewBody(BaseModel):
    """Preview parse output for client report logs."""

    log_lines: list[str]


class TaskReportParsePreviewResponse(BaseModel):
    parsed: dict[str, Any]


def _serialize_row(row: dict[str, Any]) -> dict[str, Any]:
    out = {}
    for k, v in row.items():
        if hasattr(v, "isoformat"):
            out[k] = v.isoformat() if v is not None else None
        else:
            out[k] = v
    return out


def _serialize_saved_record_row(row: dict[str, Any]) -> dict[str, Any]:
    base = _serialize_row(row)
    raw = base.get("content")
    if isinstance(raw, str) and raw.strip():
        try:
            base["content"] = json.loads(raw)
        except json.JSONDecodeError:
            base["content"] = {"_invalid_json": raw[:2000]}
    return base


def _serialize_task_row(row: dict[str, Any]) -> dict[str, Any]:
    """API task shape: core fields + params (object). Legacy DB columns are not exposed."""
    base = _serialize_row(row)
    params_obj = parse_task_params(base)
    out: dict[str, Any] = {
        "id": base["id"],
        "device_id": base.get("device_id"),
        "task_type": base["task_type"],
        "status": base["status"],
        "params": params_obj,
        "failure_detail": base.get("failure_detail"),
        "retry_count": int(base.get("retry_count") or 0),
        "created_at": base.get("created_at"),
        "updated_at": base.get("updated_at"),
        "started_at": base.get("started_at"),
        "finished_at": base.get("finished_at"),
        "environment": base.get("task_environment"),
    }
    if "device_alias" in base:
        out["device_alias"] = base["device_alias"]
    return out


def _paginate_tasks(total: int, page: int, per_page: int, rows: list) -> PaginatedTasks:
    total_pages = max(1, (total + per_page - 1) // per_page) if total else 1
    return PaginatedTasks(
        items=[_serialize_task_row(r) for r in rows],
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
    )


def _paginate_rows(total: int, page: int, per_page: int, rows: list) -> PaginatedRows:
    total_pages = max(1, (total + per_page - 1) // per_page) if total else 1
    return PaginatedRows(
        items=[_serialize_row(r) for r in rows],
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
    )


def _distribute_total(total: int, n: int) -> list[int]:
    if n <= 0:
        return []
    base = total // n
    rem = total % n
    return [base + (1 if i < rem else 0) for i in range(n)]


CLICK_TYPES = frozenset({"search_click", "related_click", "similar_click"})


def _xlsx_row_to_address(row) -> dict[str, Any] | None:
    vals = list(row) if row else []
    while len(vals) < 8:
        vals.append(None)
    if vals[0] is None and not any(vals):
        return None
    return {
        "recipient_name": str(vals[0]).strip() if vals[0] is not None else None,
        "state": (str(vals[1]).strip()[:64] if vals[1] is not None else None) or None,
        "city": (str(vals[2]).strip()[:255] if vals[2] is not None else None) or None,
        "address_line1": (str(vals[3]).strip()[:512] if vals[3] is not None else None) or None,
        "address_line2": (str(vals[4]).strip()[:512] if vals[4] is not None else None) or None,
        "zip_code": (str(vals[5]).strip()[:64] if vals[5] is not None else None) or None,
        "phone": (str(vals[6]).strip()[:64] if vals[6] is not None else None) or None,
        "full_line": (str(vals[7]).strip()[:1024] if vals[7] is not None else None) or None,
    }


app = FastAPI(title="AMZ Ops", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=os.getenv("CORS_ORIGINS", "*").split(","),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def _startup():
    db.init_tables()
    ensure_data_dir()
    try:
        db.purge_completed_tasks_older_than_days(db.get_task_retention_days())
    except Exception:
        pass


@app.post("/api/v1/auth/token", response_model=TokenResponse)
async def login(form: Annotated[OAuth2PasswordRequestForm, Depends()]):
    user = db.verify_admin_login(form.username, form.password)
    if not user:
        raise HTTPException(status_code=401, detail="鐢ㄦ埛鍚嶆垨瀵嗙爜閿欒")
    token = _create_token(user["id"], user["username"], user["is_admin"])
    return TokenResponse(access_token=token)


@app.get("/api/v1/auth/me", response_model=MeResponse)
async def me(user: CurrentUser):
    return MeResponse(id=user["id"], username=user["username"], is_admin=user["is_admin"])


@app.get("/api/v1/admin/devices/options")
async def admin_device_options(user: CurrentUser):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    rows = db.list_devices_options()
    return {"items": [_serialize_row(r) for r in rows]}


@app.get("/api/v1/admin/devices", response_model=PaginatedRows)
async def admin_devices_list(
    user: CurrentUser,
    page: int = Query(1, ge=1),
    per_page: int = Query(30, ge=1, le=100),
    q: Optional[str] = None,
):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    total, rows = db.list_devices_paginated(page, per_page, q)
    ids = [str(r["device_id"]) for r in rows]
    counts_map = db.get_pending_task_counts_by_device(ids)
    items: list[dict[str, Any]] = []
    for r in rows:
        d = _serialize_row(r)
        did = str(d["device_id"])
        d["pending_tasks"] = counts_map.get(did, {})
        items.append(d)
    total_pages = max(1, (total + per_page - 1) // per_page) if total else 1
    return PaginatedRows(
        items=items,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
    )


@app.patch("/api/v1/admin/devices/{device_id}/alias")
async def admin_device_set_alias(user: CurrentUser, device_id: str, body: DeviceAliasBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    db.set_device_alias(device_id, body.alias)
    return {"ok": True}


@app.patch("/api/v1/admin/devices/{device_id}/screenshot-upload-policy")
async def admin_device_screenshot_upload_policy(user: CurrentUser, device_id: str, body: DeviceScreenshotPolicyBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    try:
        db.upsert_device_screenshot_upload_policy(device_id, body.screenshot_upload_policy)
    except ValueError:
        raise HTTPException(status_code=400, detail="璁惧 ID 鏃犳晥")
    return {
        "ok": True,
        "screenshot_upload_policy": db.get_device_screenshot_upload_policy(device_id),
    }


@app.get("/api/v1/admin/keywords", response_model=PaginatedRows)
async def admin_keywords_list(
    user: CurrentUser,
    page: int = Query(1, ge=1),
    per_page: int = Query(30, ge=1, le=200),
    q: Optional[str] = None,
):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    total, rows = db.list_keywords_paginated(page, per_page, q)
    return _paginate_rows(total, page, per_page, rows)


@app.post("/api/v1/admin/keywords/import")
async def admin_keywords_import(user: CurrentUser, file: UploadFile = File(...)):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    raw = await file.read()
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("gbk", errors="replace")
    lines = text.splitlines()
    n = db.keywords_import_lines(lines)
    return {"ok": True, "imported": n}


@app.delete("/api/v1/admin/keywords/{keyword_id}")
async def admin_keyword_delete(user: CurrentUser, keyword_id: int):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    if not db.keyword_delete(keyword_id):
        raise HTTPException(status_code=404, detail="涓嶅瓨鍦?")
    return {"ok": True}


@app.post("/api/v1/admin/tasks/batch-click")
async def admin_tasks_batch_click(user: CurrentUser, body: BatchClickTasksBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    if body.task_type not in CLICK_TYPES:
        raise HTTPException(status_code=400, detail="task_type 鏃犳晥")
    device_ids = [d.strip() for d in body.device_ids if d and str(d).strip()]
    if not device_ids:
        raise HTTPException(status_code=400, detail="璇疯嚦灏戦€夋嫨涓€鍙拌澶?")
    pairs: list[tuple[str, int]] = []
    if body.mode == "manual":
        for d in device_ids:
            c = int(body.per_device_counts.get(d, 0))
            if c > 0:
                pairs.append((d, c))
    else:
        total = int(body.total_count)
        if total <= 0:
            raise HTTPException(status_code=400, detail="鏅鸿兘鍒嗛厤璇峰～鍐欐€讳换鍔℃暟")
        counts = _distribute_total(total, len(device_ids))
        pairs = [(device_ids[i], counts[i]) for i in range(len(device_ids)) if counts[i] > 0]
    if not pairs:
        raise HTTPException(status_code=400, detail="娌℃湁鍙垱寤虹殑浠诲姟鏁伴噺")
    fn = (body.res_folder_name or "").strip()
    if not fn:
        legacy = [str(t).strip() for t in (body.product_titles or []) if str(t).strip()]
        one = (body.product_title or "").strip()
        if legacy:
            fn = legacy[0]
        elif one:
            fn = one
    if not fn:
        raise HTTPException(status_code=400, detail="璇峰～鍐欒祫婧愭枃浠跺す鍚?res_folder_name锛堜笌鍏抽敭璇?1:1锛?")
    n = db.insert_click_tasks_batch(
        body.task_type,
        body.keyword.strip(),
        fn,
        pairs,
        persist_data=body.save_data_record,
    )
    return {"ok": True, "created": n}


class ImportLinesBody(BaseModel):
    text: str = Field("", description="澶氳鏂囨湰锛屼竴琛屼竴鏉?")


class RegisterCodePoolsStatsResponse(BaseModel):
    phone_available: int
    phone_total: int
    email_available: int
    email_total: int


class PoolDeleteBody(BaseModel):
    ids: list[int] = Field(default_factory=list)


@app.post("/api/v1/admin/tasks/batch-register")
async def admin_tasks_batch_register(user: CurrentUser, body: BatchRegisterBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    device_ids = [d.strip() for d in body.device_ids if d and str(d).strip()]
    if not device_ids:
        raise HTTPException(status_code=400, detail="璇疯嚦灏戦€夋嫨涓€鍙拌澶?")
    pairs: list[tuple[str, int]] = []
    if body.mode == "manual":
        for d in device_ids:
            c = int(body.per_device_counts.get(d, 0))
            if c > 0:
                pairs.append((d, c))
        n_tasks = sum(c for _, c in pairs)
        if n_tasks <= 0:
            raise HTTPException(status_code=400, detail="璇蜂负鎵€閫夎澶囧～鍐欏ぇ浜?0 鐨勪换鍔℃暟")
    else:
        total = int(body.total_count)
        if total <= 0:
            raise HTTPException(status_code=400, detail="鏅鸿兘鍒嗛厤璇峰～鍐欐€讳换鍔℃暟锛堜笖椤讳笉瓒呰繃鎵嬫満鎺ョ爜搴撳彲鐢ㄦ潯鏁帮級")
        counts = _distribute_total(total, len(device_ids))
        pairs = [(device_ids[i], counts[i]) for i in range(len(device_ids)) if counts[i] > 0]
        n_tasks = total
    if not pairs:
        raise HTTPException(status_code=400, detail="娌℃湁鍙垎閰嶇殑浠诲姟鏁伴噺")
    try:
        n = db.insert_register_tasks_from_pool(
            pairs, bind_email=body.bind_email, save_data_record=body.save_data_record
        )
    except ValueError as e:
        code = str(e)
        if code == "EMPTY_ADDRESS_POOL":
            raise HTTPException(status_code=400, detail="鍦板潃搴撲负绌猴紝璇峰厛瀵煎叆鍦板潃")
        if code == "REGISTER_EMPTY_DEVICE":
            raise HTTPException(status_code=400, detail="璁惧 ID 鏃犳晥")
        if code == "REGISTER_PHONE_POOL_EMPTY":
            raise HTTPException(status_code=400, detail="鎵嬫満鎺ョ爜搴撳彲鐢ㄦ潯鏁颁笉瓒筹紝璇峰厛瀵煎叆鎴栫瓑寰呬换鍔℃秷鑰?")
        if code == "REGISTER_EMAIL_POOL_EMPTY":
            raise HTTPException(status_code=400, detail="閭鎺ョ爜搴撳彲鐢ㄦ潯鏁颁笉瓒筹紙鍕鹃€夌粦瀹氶偖绠辨椂椤讳笌浠诲姟鏁颁竴鑷达級")
        raise
    if n != n_tasks:
        raise HTTPException(status_code=500, detail="鍒涘缓鏁伴噺涓庨鏈熶笉涓€鑷?")
    return {"ok": True, "created": n}


@app.get("/api/v1/admin/register-code-pools/stats", response_model=RegisterCodePoolsStatsResponse)
async def admin_register_code_pools_stats(user: CurrentUser):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    s = db.register_code_pools_stats()
    return RegisterCodePoolsStatsResponse(**s)


@app.post("/api/v1/admin/register-phone-pool/import")
async def admin_register_phone_pool_import(user: CurrentUser, body: ImportLinesBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    lines = [ln for ln in (body.text or "").splitlines()]
    n = db.import_register_phone_pool_lines(lines)
    if n == 0:
        raise HTTPException(
            status_code=400,
            detail="鏈鍏ヤ换浣曡锛岃浣跨敤鏍煎紡锛氭墜鏈哄彿----鎺ョ爜閾炬帴锛堝洓涓嫳鏂囨í绾垮垎闅旓級",
        )
    return {"ok": True, "imported": n}


@app.post("/api/v1/admin/register-email-pool/import")
async def admin_register_email_pool_import(user: CurrentUser, body: ImportLinesBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    lines = [ln for ln in (body.text or "").splitlines()]
    n = db.import_register_email_pool_lines(lines)
    if n == 0:
        raise HTTPException(
            status_code=400,
            detail="未导入任何行，请使用格式：邮箱----密码----接码地址（三段用 ---- 分隔）",
        )
    return {"ok": True, "imported": n}


@app.get("/api/v1/admin/register-phone-pool", response_model=PaginatedRows)
async def admin_register_phone_pool_list(
    user: CurrentUser,
    page: int = Query(1, ge=1),
    per_page: int = Query(30, ge=1, le=100),
    q: Optional[str] = None,
):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    total, rows = db.list_register_phone_pool_paginated(page, per_page, q)
    return _paginate_rows(total, page, per_page, rows)


@app.post("/api/v1/admin/register-phone-pool/delete")
async def admin_register_phone_pool_delete(user: CurrentUser, body: PoolDeleteBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    ids = [int(x) for x in body.ids if int(x) > 0]
    if not ids:
        raise HTTPException(status_code=400, detail="璇烽€夋嫨瑕佸垹闄ょ殑璁板綍")
    n = db.delete_register_phone_pool_ids(ids)
    return {"ok": True, "deleted": n}


@app.post("/api/v1/admin/register-phone-pool/clear")
async def admin_register_phone_pool_clear(user: CurrentUser):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    n = db.clear_register_phone_pool()
    return {"ok": True, "deleted": n}


@app.get("/api/v1/admin/register-email-pool", response_model=PaginatedRows)
async def admin_register_email_pool_list(
    user: CurrentUser,
    page: int = Query(1, ge=1),
    per_page: int = Query(30, ge=1, le=100),
    q: Optional[str] = None,
):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    total, rows = db.list_register_email_pool_paginated(page, per_page, q)
    out = []
    for r in rows:
        item = _serialize_row(r)
        pw = item.get("email_login_password")
        if isinstance(pw, str) and len(pw) > 4:
            item["email_login_password_masked"] = pw[:2] + "..." + pw[-2:]
        else:
            item["email_login_password_masked"] = "****"
        item.pop("email_login_password", None)
        out.append(item)
    total_pages = max(1, (total + per_page - 1) // per_page) if total else 1
    return PaginatedRows(
        items=out,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
    )


@app.post("/api/v1/admin/register-email-pool/delete")
async def admin_register_email_pool_delete(user: CurrentUser, body: PoolDeleteBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    ids = [int(x) for x in body.ids if int(x) > 0]
    if not ids:
        raise HTTPException(status_code=400, detail="璇烽€夋嫨瑕佸垹闄ょ殑璁板綍")
    n = db.delete_register_email_pool_ids(ids)
    return {"ok": True, "deleted": n}


@app.post("/api/v1/admin/register-email-pool/clear")
async def admin_register_email_pool_clear(user: CurrentUser):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    n = db.clear_register_email_pool()
    return {"ok": True, "deleted": n}


@app.get("/api/v1/admin/settings", response_model=AdminSettingsResponse)
async def admin_get_settings(user: CurrentUser):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    return AdminSettingsResponse(task_retention_days=db.get_task_retention_days())


@app.patch("/api/v1/admin/settings", response_model=AdminSettingsResponse)
async def admin_patch_settings(user: CurrentUser, body: AdminSettingsPatchBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    db.set_task_retention_days(body.task_retention_days)
    try:
        db.purge_completed_tasks_older_than_days(db.get_task_retention_days())
    except Exception:
        pass
    return AdminSettingsResponse(task_retention_days=db.get_task_retention_days())


@app.get("/api/v1/admin/task-saved-records", response_model=PaginatedRows)
async def admin_task_saved_records_list(
    user: CurrentUser,
    page: int = Query(1, ge=1),
    per_page: int = Query(30, ge=1, le=100),
    task_type: Optional[str] = None,
    q: Optional[str] = None,
):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    total, rows = db.list_task_saved_records(page, per_page, task_type, q)
    total_pages = max(1, (total + per_page - 1) // per_page) if total else 1
    return PaginatedRows(
        items=[_serialize_saved_record_row(r) for r in rows],
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
    )


@app.get("/api/v1/admin/addresses", response_model=PaginatedRows)
async def admin_addresses_list(
    user: CurrentUser,
    page: int = Query(1, ge=1),
    per_page: int = Query(30, ge=1, le=100),
    q: Optional[str] = None,
):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    total, rows = db.list_addresses_paginated(page, per_page, q)
    return _paginate_rows(total, page, per_page, rows)


@app.post("/api/v1/admin/addresses")
async def admin_address_create(user: CurrentUser, body: UsAddressBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    aid = db.address_create(body.model_dump())
    return {"ok": True, "id": aid}


@app.get("/api/v1/admin/addresses/{address_id}")
async def admin_address_get(user: CurrentUser, address_id: int):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    row = db.address_get(address_id)
    if not row:
        raise HTTPException(status_code=404, detail="涓嶅瓨鍦?")
    return _serialize_row(row)


@app.put("/api/v1/admin/addresses/{address_id}")
async def admin_address_update(user: CurrentUser, address_id: int, body: UsAddressBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    if not db.address_update(address_id, body.model_dump()):
        raise HTTPException(status_code=404, detail="涓嶅瓨鍦?")
    return {"ok": True}


@app.delete("/api/v1/admin/addresses/{address_id}")
async def admin_address_delete(user: CurrentUser, address_id: int):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    if not db.address_delete(address_id):
        raise HTTPException(status_code=404, detail="涓嶅瓨鍦?")
    return {"ok": True}


@app.post("/api/v1/admin/addresses/import-xlsx")
async def admin_addresses_import_xlsx(user: CurrentUser, file: UploadFile = File(...)):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="鏈嶅姟绔湭瀹夎 openpyxl")
    data = await file.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_out: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        rec = _xlsx_row_to_address(row)
        if rec and (rec.get("recipient_name") or rec.get("full_line") or rec.get("address_line1")):
            rows_out.append(rec)
    wb.close()
    if not rows_out:
        raise HTTPException(status_code=400, detail="鏈В鏋愬埌鏈夋晥琛?")
    n = db.addresses_import_rows(rows_out)
    return {"ok": True, "imported": n}


@app.get("/api/v1/admin/target-asins", response_model=PaginatedRows)
async def admin_target_asins_list(
    user: CurrentUser,
    page: int = Query(1, ge=1),
    per_page: int = Query(30, ge=1, le=100),
    q: Optional[str] = None,
):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    total, rows = db.list_target_asins_paginated(page, per_page, q)
    return _paginate_rows(total, page, per_page, rows)


@app.post("/api/v1/admin/target-asins")
async def admin_target_asin_create(user: CurrentUser, body: TargetAsinCreateBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    try:
        aid = db.target_asin_create(body.asin, body.note or None)
    except ValueError:
        raise HTTPException(status_code=400, detail="ASIN 鏍煎紡鏃犳晥锛堥渶 10锝?6 浣嶅瓧姣嶆暟瀛楋級")
    except pymysql.err.IntegrityError:
        raise HTTPException(status_code=400, detail="ASIN 宸插瓨鍦?")
    return {"ok": True, "id": aid}


@app.get("/api/v1/admin/target-asins/{asin_id}")
async def admin_target_asin_get(user: CurrentUser, asin_id: int):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    row = db.target_asin_get(asin_id)
    if not row:
        raise HTTPException(status_code=404, detail="涓嶅瓨鍦?")
    return _serialize_row(row)


@app.put("/api/v1/admin/target-asins/{asin_id}")
async def admin_target_asin_update(user: CurrentUser, asin_id: int, body: TargetAsinPatchBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    if body.asin is None and body.note is None:
        raise HTTPException(status_code=400, detail="璇疯嚦灏戞彁渚涗竴涓淇敼鐨勫瓧娈?")
    try:
        ok = db.target_asin_update(asin_id, body.asin, body.note)
    except ValueError:
        raise HTTPException(status_code=400, detail="ASIN 鏍煎紡鏃犳晥锛堥渶 10锝?6 浣嶅瓧姣嶆暟瀛楋級")
    except pymysql.err.IntegrityError:
        raise HTTPException(status_code=400, detail="ASIN 涓庡叾浠栬褰曞啿绐?")
    if not ok:
        raise HTTPException(status_code=404, detail="涓嶅瓨鍦?")
    return {"ok": True}


@app.delete("/api/v1/admin/target-asins/{asin_id}")
async def admin_target_asin_delete(user: CurrentUser, asin_id: int):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    if not db.target_asin_delete(asin_id):
        raise HTTPException(status_code=404, detail="涓嶅瓨鍦?")
    return {"ok": True}


@app.get("/api/v1/admin/asin-click-records", response_model=PaginatedRows)
async def admin_asin_click_records_list(
    user: CurrentUser,
    page: int = Query(1, ge=1),
    per_page: int = Query(40, ge=1, le=100),
    q: Optional[str] = None,
    asin: Optional[str] = Query(None, description="鎸?ASIN 绮剧‘绛涢€夛紙涓庣洰鏍?ASIN 琛ㄤ竴鑷寸殑澶у啓瑙勮寖褰㈠紡锛?"),
):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    total, rows = db.list_asin_click_records_paginated(page, per_page, q, asin)
    return _paginate_rows(total, page, per_page, rows)


@app.post("/api/v1/admin/task-report/parse-preview", response_model=TaskReportParsePreviewResponse)
async def admin_task_report_parse_preview(user: CurrentUser, body: TaskReportParsePreviewBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    p = parse_task_report_footer(body.log_lines)
    fin: str | None = None
    if p.finished_at is not None:
        fin = p.finished_at.isoformat()
    return TaskReportParsePreviewResponse(
        parsed={
            "success": p.success,
            "environment": p.environment,
            "finished_at": fin,
            "failure_detail": p.failure_detail,
            "used_amz_report": p.used_amz_report,
        }
    )


@app.get("/api/v1/admin/task-center/tasks", response_model=PaginatedTasks)
async def admin_task_center_list(
    user: CurrentUser,
    page: int = Query(1, ge=1),
    per_page: int = Query(30, ge=1, le=100),
    device_id: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    task_type: Optional[str] = None,
    params_q: Optional[str] = Query(
        None,
        max_length=256,
        description="鍦?params JSON 鍙婇仐鐣欏叧閿瘝/鏍囬/鎵嬫満绛夊瓧娈典腑妯＄硦鍖归厤",
    ),
):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    pq = params_q.strip() if params_q and params_q.strip() else None
    total, rows = db.list_tasks_filtered(page, per_page, device_id, status_filter, task_type, pq)
    return _paginate_tasks(total, page, per_page, rows)


@app.delete("/api/v1/admin/task-center/tasks")
async def admin_task_delete_all_matching(
    user: CurrentUser,
    device_id: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    task_type: Optional[str] = None,
    params_q: Optional[str] = Query(
        None,
        max_length=256,
        description="涓庡垪琛ㄦ帴鍙ｄ竴鑷达紱鏃犵瓫閫夊弬鏁版椂鍒犻櫎搴撳唴鍏ㄩ儴浠诲姟",
    ),
):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    pq = params_q.strip() if params_q and params_q.strip() else None
    deleted, names = db.delete_tasks_matching_filters(device_id, status_filter, task_type, pq)
    for name in names:
        p = os.path.join(TASK_IMAGE_DIR, name)
        try:
            os.remove(p)
        except FileNotFoundError:
            pass
    return {"ok": True, "deleted": deleted}


@app.get("/api/v1/admin/task-center/tasks/{task_id}")
async def admin_task_center_detail(user: CurrentUser, task_id: int):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    task = db.get_task_by_id(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="浠诲姟涓嶅瓨鍦?")
    logs = db.get_task_logs(task_id)
    imgs = db.get_task_image_rows(task_id)
    return {
        "task": _serialize_task_row(task),
        "logs": [_serialize_row(r) for r in logs],
        "screenshots": [_serialize_row(r) for r in imgs],
    }


@app.get("/api/v1/admin/task-center/screenshots/{image_id}")
async def admin_task_screenshot_file(user: CurrentUser, image_id: int):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    row = db.get_image_by_id(image_id)
    if not row:
        raise HTTPException(status_code=404, detail="涓嶅瓨鍦?")
    path = os.path.join(TASK_IMAGE_DIR, row["stored_name"])
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="鏂囦欢缂哄け")
    media_type, _ = mimetypes.guess_type(path)
    return FileResponse(path, media_type=media_type or "application/octet-stream")


@app.post("/api/v1/admin/task-center/tasks/{task_id}/retry")
async def admin_task_retry(user: CurrentUser, task_id: int):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    ok, names = db.retry_failed_task(task_id)
    if not ok:
        raise HTTPException(status_code=400, detail="浠呭け璐ョ姸鎬佸彲閲嶈瘯")
    for name in names:
        p = os.path.join(TASK_IMAGE_DIR, name)
        try:
            os.remove(p)
        except FileNotFoundError:
            pass
    return {"ok": True}


@app.delete("/api/v1/admin/task-center/tasks/{task_id}")
async def admin_task_delete(user: CurrentUser, task_id: int):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    ok, names = db.delete_task_and_collect_images(task_id)
    if not ok:
        raise HTTPException(status_code=404, detail="浠诲姟涓嶅瓨鍦?")
    for name in names:
        p = os.path.join(TASK_IMAGE_DIR, name)
        try:
            os.remove(p)
        except FileNotFoundError:
            pass
    return {"ok": True}


@app.post("/api/v1/admin/task-center/tasks/{task_id}/redo")
async def admin_task_redo(user: CurrentUser, task_id: int):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    try:
        new_id = db.clone_task_redo(task_id)
    except ValueError as e:
        if str(e) == "EMPTY_ADDRESS_POOL":
            raise HTTPException(status_code=400, detail="鍦板潃搴撲负绌猴紝鏃犳硶澶嶅埗娉ㄥ唽浠诲姟")
        raise
    if not new_id:
        raise HTTPException(status_code=400, detail="鏃犳硶澶嶅埗璇ヤ换鍔★紙绫诲瀷涓嶆敮鎸佹垨鏁版嵁涓嶅畬鏁达級")
    return {"ok": True, "new_task_id": new_id}


@app.post("/api/v1/client/heartbeat")
async def client_heartbeat(body: HeartbeatBody):
    db.upsert_device_heartbeat(body.device_id)
    pol = db.get_device_screenshot_upload_policy(body.device_id)
    return {"ok": True, "screenshot_upload_policy": pol}


@app.post("/api/v1/client/amazon-accounts/bootstrap")
async def client_amazon_accounts_bootstrap(body: ClientAmazonBootstrapBody):
    """Create/update amazon account record after register success."""
    db.upsert_device_heartbeat(body.device_id)
    row = db.amazon_account_bootstrap(body.task_id, body.device_id, None)
    if not row:
        raise HTTPException(status_code=400, detail="浠诲姟涓嶅瓨鍦ㄣ€侀潪 register 绫诲瀷鎴栬澶囦笉鍖归厤")
    out = _serialize_row(row)
    if "totp_secret" in out:
        del out["totp_secret"]
    return {"ok": True, "account": out}


@app.post("/api/v1/client/amazon-accounts/totp-done")
async def client_amazon_accounts_totp_done(body: ClientAmazonAccountStageBody):
    ok = db.amazon_account_mark_totp_set_by_phone(body.phone)
    if not ok:
        raise HTTPException(status_code=400, detail="手机号码未匹配到账号记录")
    return {"ok": True}


@app.post("/api/v1/client/amazon-accounts/address-done")
async def client_amazon_accounts_address_done(body: ClientAmazonAccountStageBody):
    ok = db.amazon_account_mark_address_set_by_phone(body.phone)
    if not ok:
        raise HTTPException(status_code=400, detail="手机号码未匹配到账号记录")
    return {"ok": True}

@app.post("/api/v1/client/amazon-accounts/totp-qr")
async def client_amazon_accounts_totp_qr(
    phone: str = Form(..., min_length=1, max_length=64),
    device_id: str = Form(...),
    image: UploadFile = File(...),
):
    db.upsert_device_heartbeat(device_id)
    raw = await image.read()
    if not raw:
        raise HTTPException(status_code=400, detail="鍥剧墖涓虹┖")
    out = db.amazon_account_apply_totp_upload(
        phone, raw, image.filename or "totp.png"
    )
    if not out.get("ok"):
        raise HTTPException(status_code=400, detail=out.get("error") or "澶勭悊澶辫触")
    return {
        "ok": True,
        "totp_code": out.get("totp_code"),
        "secret_saved": bool(out.get("secret_saved")),
    }


@app.get("/api/v1/client/amazon-accounts/totp-code")
async def client_amazon_accounts_totp_code(
    phone: str = Query(..., min_length=1, max_length=64),
    device_id: Optional[str] = Query(None),
):
    did = (device_id or "").strip()
    if did:
        db.upsert_device_heartbeat(did)
    data = db.amazon_account_totp_code_by_phone(phone)
    if data is None:
        raise HTTPException(status_code=400, detail="手机号码未匹配到账号记录")
    return {"ok": True, **data}


@app.get("/api/v1/admin/amazon-accounts", response_model=PaginatedRows)
async def admin_amazon_accounts_list(
    user: CurrentUser,
    page: int = Query(1, ge=1),
    per_page: int = Query(30, ge=1, le=100),
    q: Optional[str] = None,
):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    total, rows = db.list_amazon_accounts_paginated(page, per_page, q)
    items: list[dict[str, Any]] = []
    for r in rows:
        d = _serialize_row(r)
        sec = r.get("totp_secret")
        d["totp_code_now"] = totp_current_code(str(sec)) if sec else None
        d["totp_image_url"] = f"/api/v1/admin/amazon-accounts/{int(r['id'])}/totp-image"
        if "env_name" not in d and "environment" in d:
            d["env_name"] = d.get("environment")
        d["totp_configured"] = bool(r.get("totp_set_at")) or bool(sec)
        d["address_configured"] = bool(r.get("address_set_at"))
        d.pop("totp_secret", None)
        items.append(d)
    total_pages = max(1, (total + per_page - 1) // per_page) if total else 1
    return PaginatedRows(
        items=items,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
    )


@app.get("/api/v1/admin/amazon-accounts/{account_id}/totp-image")
async def admin_amazon_account_totp_image(user: CurrentUser, account_id: int):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    row = db.get_amazon_account_by_id(account_id)
    if not row:
        raise HTTPException(status_code=404, detail="璁板綍涓嶅瓨鍦?")
    stored_name = row.get("totp_image_stored_name")
    path = safe_task_image_path(str(stored_name) if stored_name else None)
    if not path or not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="TOTP鍥剧墖涓嶅瓨鍦?")
    media_type = mimetypes.guess_type(path)[0] or "application/octet-stream"
    return FileResponse(path, media_type=media_type)


@app.delete("/api/v1/admin/amazon-accounts/{account_id}")
async def admin_amazon_account_delete(user: CurrentUser, account_id: int):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    ok = db.delete_amazon_account_by_id(account_id)
    if not ok:
        raise HTTPException(status_code=404, detail="璁板綍涓嶅瓨鍦?")
    return {"ok": True}


@app.post("/api/v1/admin/amazon-accounts/delete")
async def admin_amazon_accounts_delete(user: CurrentUser, body: PoolDeleteBody):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    ids = [int(x) for x in body.ids if int(x) > 0]
    if not ids:
        raise HTTPException(status_code=400, detail="璇烽€夋嫨瑕佸垹闄ょ殑璁板綍")
    n = db.delete_amazon_accounts_ids(ids)
    return {"ok": True, "deleted": n}


@app.post("/api/v1/admin/amazon-accounts/clear")
async def admin_amazon_accounts_clear(user: CurrentUser):
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="闇€瑕佺鐞嗗憳鏉冮檺")
    n = db.clear_amazon_accounts()
    return {"ok": True, "deleted": n}


def _captcha_image_dimensions(body: bytes) -> tuple[int, int]:
    try:
        from PIL import Image

        im = Image.open(io.BytesIO(body))
        return int(im.width), int(im.height)
    except Exception:
        return 0, 0


@app.post("/api/v1/client/captcha-assist/upload")
async def client_captcha_assist_upload(
    task_id: int = Form(...),
    device_id: str = Form(...),
    image: UploadFile = File(...),
):
    """Client uploads captcha screenshot for manual assist."""
    db.upsert_device_heartbeat(device_id)
    did = device_id.strip()
    t = db.get_task_by_id(int(task_id))
    if not t:
        raise HTTPException(status_code=404, detail="浠诲姟涓嶅瓨鍦?")
    if str(t.get("device_id") or "").strip() != did:
        raise HTTPException(status_code=403, detail="璁惧涓嶅尮閰?")
    if (t.get("status") or "") != "running":
        raise HTTPException(status_code=400, detail="浠呮墽琛屼腑浠诲姟鍙笂浼犱汉宸ラ獙璇佺爜鎴浘")
    body = await image.read()
    if not body:
        raise HTTPException(status_code=400, detail="鍥剧墖涓虹┖")
    max_mb = int(os.getenv("TASK_IMAGE_MAX_MB", "8"))
    if len(body) > max_mb * 1024 * 1024:
        raise HTTPException(status_code=413, detail="鍥剧墖杩囧ぇ")
    w, h = _captcha_image_dimensions(body)
    ext = Path(image.filename or "").suffix.lower()
    if ext not in (".jpg", ".jpeg", ".png", ".webp"):
        ext = ".jpg"
    stored = f"captcha_{int(task_id)}_{uuid.uuid4().hex}{ext}"
    ensure_data_dir()
    dest = os.path.join(TASK_IMAGE_DIR, stored)
    with open(dest, "wb") as f:
        f.write(body)
    sid = db.captcha_assist_create(int(task_id), did, stored, w, h)
    if not sid:
        try:
            os.remove(dest)
        except OSError:
            pass
        raise HTTPException(status_code=500, detail="鍒涘缓浼氳瘽澶辫触")
    return {"ok": True, "session_id": sid, "image_width": w, "image_height": h}


@app.get("/api/v1/client/captcha-assist/result")
async def client_captcha_assist_result(
    task_id: int = Query(..., ge=1),
    device_id: str = Query(..., min_length=1),
    session_id: int = Query(..., ge=1),
):
    db.upsert_device_heartbeat(device_id)
    out = db.captcha_assist_client_poll(task_id, device_id, session_id)
    if not out:
        raise HTTPException(status_code=403, detail="绂佹璁块棶")
    return {"ok": True, **out}


class CaptchaAssistSubmitBody(BaseModel):
    clicks: list[dict[str, Any]] = Field(..., min_length=1)


@app.get("/api/v1/admin/captcha-assist/pending")
async def admin_captcha_assist_pending():
    """List pending captcha assist sessions."""
    rows = db.captcha_assist_list_pending()
    return {"items": [_serialize_row(r) for r in rows]}


@app.get("/api/v1/admin/captcha-assist/sessions/{session_id}/image")
async def admin_captcha_assist_image(session_id: int):
    """Fetch captcha assist session image."""
    row = db.captcha_assist_get(session_id)
    if not row:
        raise HTTPException(status_code=404, detail="浼氳瘽涓嶅瓨鍦?")
    name = row.get("image_stored_name") or ""
    path = safe_task_image_path(str(name))
    if not path or not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="鍥剧墖涓嶅瓨鍦?")
    mt = mimetypes.guess_type(name)[0] or "image/jpeg"
    return FileResponse(path, media_type=mt)


@app.post("/api/v1/admin/captcha-assist/sessions/{session_id}/submit")
async def admin_captcha_assist_submit(
    session_id: int,
    body: CaptchaAssistSubmitBody,
):
    clicks: list[dict[str, int]] = []
    for c in body.clicks:
        if not isinstance(c, dict):
            continue
        try:
            x = int(float(c.get("x")))
            y = int(float(c.get("y")))
        except (TypeError, ValueError):
            continue
        clicks.append({"x": x, "y": y})
    if not clicks:
        raise HTTPException(status_code=400, detail="鍧愭爣鏃犳晥")
    ok = db.captcha_assist_submit_clicks(session_id, clicks)
    if not ok:
        raise HTTPException(status_code=400, detail="鎻愪氦澶辫触锛堜細璇濅笉瀛樺湪鎴栧凡澶勭悊锛?")
    return {"ok": True}


@app.post("/api/v1/client/asin-clicks")
async def client_asin_click(body: ClientAsinClickBody):
    """Record one target ASIN click from client."""
    db.upsert_device_heartbeat(body.device_id)
    norm = normalize_target_asin(body.asin)
    if not norm:
        raise HTTPException(status_code=400, detail="ASIN 鏍煎紡鏃犳晥锛堥渶 10锝?6 浣嶅瓧姣嶆暟瀛楋級")
    kw = (body.keyword or "").strip()
    if not kw:
        raise HTTPException(status_code=400, detail="keyword 涓嶈兘涓虹┖")
    out = db.increment_target_asin_click(norm, kw, body.device_id)
    if not out:
        raise HTTPException(status_code=500, detail="璁板綍鐐瑰嚮澶辫触")
    return {"ok": True, **out}


@app.get("/api/v1/client/random-keywords")
async def client_random_keywords(num: int = Query(2, ge=1, le=100)):
    kws = db.keywords_random_sample(num)
    return {"keywords": kws, "count": len(kws)}


@app.get("/api/v1/client/tasks/next")
async def client_tasks_next(
    device_id: str = Query(..., min_length=1),
    task_type: Optional[str] = Query(None, description="浠呴鍙栨寚瀹氱被鍨嬶紝濡?search_click"),
):
    db.upsert_device_heartbeat(device_id)
    pol = db.get_device_screenshot_upload_policy(device_id)
    tt = task_type.strip() if task_type else None
    task = db.claim_next_task(device_id, tt)
    if not task:
        return {"task": None, "screenshot_upload_policy": pol}
    return {"task": _serialize_task_row(task), "screenshot_upload_policy": pol}


@app.post("/api/v1/client/tasks/{task_id}/screenshots")
async def client_task_append_screenshots(
    task_id: int,
    device_id: str = Form(...),
    description: str = Form(default=""),
    image: UploadFile = File(...),
):
    """Upload one running/failed task screenshot according to policy."""
    task = db.get_task_by_id(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="浠诲姟涓嶅瓨鍦?")
    if str(task.get("device_id") or "") != device_id.strip():
        raise HTTPException(status_code=403, detail="璁惧涓嶅尮閰?")
    policy = db.get_device_screenshot_upload_policy(device_id)
    st = task["status"]
    if policy == "none":
        raise HTTPException(status_code=403, detail="璁惧绛栫暐涓虹姝笂浼犳埅鍥?")
    if st == "running":
        if policy == "failed_only":
            raise HTTPException(
                status_code=403,
                detail="设备策略为 failed_only 时，运行中请先本地暂存截图，任务失败结案后再补传。",
            )
    elif st == "failed":
        if policy not in ("all", "failed_only"):
            raise HTTPException(status_code=403, detail="璁惧绛栫暐涓虹姝笂浼犳埅鍥?")
    else:
        raise HTTPException(status_code=400, detail="浠呮墽琛屼腑鎴栧凡澶辫触浠诲姟鍙笂浼犺繃绋嬫埅鍥?")
    if not image.filename:
        raise HTTPException(status_code=400, detail="缂哄皯鍥剧墖鏂囦欢")
    body = await image.read()
    max_mb = int(os.getenv("TASK_IMAGE_MAX_MB", "8"))
    if len(body) > max_mb * 1024 * 1024:
        raise HTTPException(status_code=413, detail="鍗曞紶鍥剧墖杩囧ぇ")
    ext = Path(image.filename).suffix.lower() or ".bin"
    if ext not in (".png", ".jpg", ".jpeg", ".webp", ".gif", ".bin"):
        ext = ".bin"
    stored = f"{task_id}_{uuid.uuid4().hex}{ext}"
    dest = os.path.join(TASK_IMAGE_DIR, stored)
    with open(dest, "wb") as f:
        f.write(body)
    desc = (description or "").strip()[:512] or None
    img_id = db.insert_task_image(task_id, stored, desc)
    return {"ok": True, "image_id": img_id}


@app.post("/api/v1/client/tasks/{task_id}/report")
async def client_task_report(
    task_id: int,
    device_id: str = Form(...),
    log_lines: str = Form(default="[]"),
    images: Optional[list[UploadFile]] = File(None),
    image_descriptions: Optional[str] = Form(
        default=None,
        description="可选，与 images 顺序一致的说明 JSON 数组，例如 [\"search\",\"detail\"]",
    ),
):
    try:
        lines = json.loads(log_lines) if log_lines else []
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="log_lines 椤讳负 JSON 瀛楃涓叉暟缁?")
    if not isinstance(lines, list):
        raise HTTPException(status_code=400, detail="log_lines 椤讳负鏁扮粍")
    lines = [str(x) for x in lines]
    task = db.get_task_by_id(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="浠诲姟涓嶅瓨鍦?")
    if task["status"] != "running":
        raise HTTPException(status_code=400, detail="浠诲姟鐘舵€佷笉鍙笂鎶?")
    if str(task.get("device_id") or "") != device_id.strip():
        raise HTTPException(status_code=403, detail="璁惧涓嶅尮閰?")
    policy = db.get_device_screenshot_upload_policy(device_id)
    parsed_preview = parse_task_report_footer(lines)
    imgs: list[UploadFile] = list(images or [])
    if policy == "none" and imgs:
        raise HTTPException(status_code=400, detail="璁惧绛栫暐涓虹姝笂浼犳埅鍥撅紝缁撴璇峰嬁闄勫甫鍥剧墖")
    if policy == "failed_only" and parsed_preview.success:
        imgs = []
    if lines:
        db.append_task_logs(task_id, lines)
    desc_list: list[str | None] = []
    if image_descriptions and str(image_descriptions).strip():
        try:
            arr = json.loads(str(image_descriptions).strip())
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="image_descriptions 椤讳负 JSON 鏁扮粍瀛楃涓?")
        if not isinstance(arr, list):
            raise HTTPException(status_code=400, detail="image_descriptions 椤讳负鏁扮粍")
        for x in arr:
            if x is None:
                desc_list.append(None)
            else:
                s = str(x).strip()
                desc_list.append(s[:512] if s else None)
    max_mb = int(os.getenv("TASK_IMAGE_MAX_MB", "8"))
    for i, uf in enumerate(imgs):
        if not uf.filename:
            continue
        body = await uf.read()
        if len(body) > max_mb * 1024 * 1024:
            raise HTTPException(status_code=413, detail="鍗曞紶鍥剧墖杩囧ぇ")
        ext = Path(uf.filename).suffix.lower() or ".bin"
        if ext not in (".png", ".jpg", ".jpeg", ".webp", ".gif", ".bin"):
            ext = ".bin"
        stored = f"{task_id}_{uuid.uuid4().hex}{ext}"
        dest = os.path.join(TASK_IMAGE_DIR, stored)
        with open(dest, "wb") as f:
            f.write(body)
        one_desc = desc_list[i] if i < len(desc_list) else None
        db.insert_task_image(task_id, stored, one_desc)
    ok, _ = db.finalize_task_from_report(task_id, lines)
    if not ok:
        raise HTTPException(status_code=400, detail="鏃犳硶缁撴浠诲姟")
    return {"ok": True}


_assets = os.path.join(STATIC, "assets")
if os.path.isdir(_assets):
    app.mount("/static/assets", StaticFiles(directory=_assets), name="static_assets")


@app.get("/static/favicon.svg")
async def admin_favicon():
    p = os.path.join(STATIC, "favicon.svg")
    if os.path.isfile(p):
        return FileResponse(p)
    raise HTTPException(status_code=404)


@app.get("/")
async def spa_index():
    index = os.path.join(STATIC, "index.html")
    if not os.path.isfile(index):
        return {
            "detail": "请先构建前端：cd admin && npm run build",
        }
    return FileResponse(index)


@app.get("/{full_path:path}")
async def spa_fallback(full_path: str):
    if full_path == "api" or full_path.startswith("api/"):
        raise HTTPException(status_code=404, detail="Not found")
    if full_path == "static" or full_path.startswith("static/"):
        raise HTTPException(status_code=404, detail="Not found")
    candidate = os.path.join(STATIC, full_path)
    if os.path.isfile(candidate):
        return FileResponse(candidate)
    index = os.path.join(STATIC, "index.html")
    if os.path.isfile(index):
        return FileResponse(index)
    raise HTTPException(status_code=404, detail="Not found")



